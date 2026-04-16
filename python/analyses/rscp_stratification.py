r"""
Regional, gender and sectoral wage stratification using ISPV / RSCP data.

ISPV (*Informační systém o průměrném výdělku*) semi-annual Excel workbooks
publish wage breakdowns by NACE sector, Czech region (kraj / NUTS3), and sex,
with full percentile profiles (P10–P90).  This script extracts three
argumentation figures that together illustrate the multi-dimensional structure
of wage inequality in the Czech labour market.

Figure A – ``rscp_regional_wages``
    Horizontal bar chart: CZ median monthly wage by region (kraj), indexed
    to the national median (100 = ČR celkem).

    Data: ISPV Excel regional sheet (``sheet_name`` matching "kraj" or index 1)
    when available; falls back to Eurostat ``earn_rgnmhw`` (mean hourly wages,
    NUTS2) for AT, DE, DK, PL, SK cross-country context.

    Argumentation: Persistent regional wage gaps (Prague 140–150 vs. rural
    regions 80–90) illustrate why a single national wage floor / collective
    agreement can mean very different things across regions — and why regional
    extension mechanisms matter for equality.

Figure B – ``rscp_gender_gap``
    Dual panel:
      Left  – CZ unadjusted gender pay gap (%) by NACE sector (ISPV M/Ž
              median columns or Eurostat ``earn_gr_gpgr2``).
      Right – Cross-country unadjusted GPG trend 2010–2024 for 6 countries
              (Eurostat ``earn_gr_gpgr2``, economy-wide).

    Argumentation: The sectoral GPG panel shows that the gender wage gap is
    not uniform — it is smallest in sectors with strong collective agreements
    (industry, finance) and largest in mixed/service sectors.  The
    cross-country trend shows CZ consistently above the EU average, providing
    a reform motivation.

Figure C – ``rscp_sector_percentiles``
    Grouped horizontal bar (floating whisker) chart: P25 / P50 / P75 wage
    levels by NACE sector for CZ (ISPV percentile columns).  Bars run from
    P25 to P75; tick marks at P50.

    Argumentation: Sectors with narrow P25–P75 bands (low within-sector
    dispersion) are candidates for sector-wide collective agreements with
    tight wage grids; wide-band sectors have heterogeneous workforces where
    economy-wide KS floors are less effective.

Data sources
------------
Czech sector/regional wages: MPSV/TREXIMA ISPV & RSCP Excel workbooks.
Gender pay gap by NACE:      Eurostat ``earn_gr_gpgr2``.
Regional hourly wages:       Eurostat ``earn_rgnmhw``.

Output
------
  pics/python/rscp_regional_wages.pdf
  pics/python/rscp_gender_gap.pdf
  pics/python/rscp_sector_percentiles.pdf
  latex/texparts/python/rscp_regional_wages.tex
  latex/texparts/python/rscp_gender_gap.tex
  latex/texparts/python/rscp_sector_percentiles.tex

Run
---
    python analyses/rscp_stratification.py
"""

from __future__ import annotations

import logging
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).parent.parent))

import matplotlib.patches as mpatches
import matplotlib.pyplot as plt
import matplotlib.ticker as ticker
import numpy as np
import pandas as pd

from config import COUNTRY_COLORS, FONT_SIZE, LATEX_PICS_DIR, PALETTE
from stattool.fetch import fetch, fetch_ispv, fetch_eurostat
from stattool.style import apply_style, cm2in, savefig, save_figure_tex

logging.basicConfig(level=logging.WARNING)
log = logging.getLogger(__name__)

apply_style()

# ── Parameters ────────────────────────────────────────────────────────────────
COUNTRIES = ["CZ", "AT", "DE", "DK", "PL", "SK"]
START_YEAR = 2010   # gender pay gap trend starts here
END_YEAR   = 2024

CZ_COLOR = COUNTRY_COLORS["CZ"]

# Czech regions (kraje) — canonical order (Prague first, then Bohemia, Moravia)
CZ_REGIONS_ORDER = [
    "Hlavní město Praha",
    "Jihomoravský kraj",
    "Plzeňský kraj",
    "Středočeský kraj",
    "Liberecký kraj",
    "Královéhradecký kraj",
    "Pardubický kraj",
    "Kraj Vysočina",
    "Jihočeský kraj",
    "Olomoucký kraj",
    "Zlínský kraj",
    "Moravskoslezský kraj",
    "Ústecký kraj",
    "Karlovarský kraj",
]

# Short region labels for axis ticks
_REGION_SHORT: dict[str, str] = {
    "Hlavní město Praha":    "Praha",
    "Středočeský kraj":      "Středočeský",
    "Jihočeský kraj":        "Jihočeský",
    "Plzeňský kraj":         "Plzeňský",
    "Karlovarský kraj":      "Karlovarský",
    "Ústecký kraj":          "Ústecký",
    "Liberecký kraj":        "Liberecký",
    "Královéhradecký kraj":  "Královéhradecký",
    "Pardubický kraj":       "Pardubický",
    "Kraj Vysočina":         "Vysočina",
    "Jihomoravský kraj":     "Jihomoravský",
    "Olomoucký kraj":        "Olomoucký",
    "Zlínský kraj":          "Zlínský",
    "Moravskoslezský kraj":  "Moravskoslezský",
}

# ── ISPV 2025 GUID URLs for 14 regional workbooks (NUTS3) ────────────────
_ISPV_REGIONAL_GUIDS: dict[str, str] = {
    "CZ010": "4da33343-6956-4838-9ced-4e4f7300e3fd",
    "CZ020": "096ded92-bf14-4434-9d0a-59cbc4c277c1",
    "CZ031": "9fdc948b-bf4d-481a-9f11-4ced6b11f627",
    "CZ032": "fd36f92f-631b-4913-b69a-0559f627b0f9",
    "CZ041": "612f031e-94e9-44fd-88ec-01b8d38a36e8",
    "CZ042": "51efc891-2c04-4891-9f02-fb43593b1417",
    "CZ051": "218633e1-9fe3-4772-8a96-fb76f17b55ce",
    "CZ052": "2c5b808e-8c20-4a79-901f-5e954fe62ac1",
    "CZ053": "5e34396a-d607-43cf-9077-5bdc14d23381",
    "CZ063": "61b8fdf4-45ee-4a06-afe7-91adbeab5ed8",
    "CZ064": "3c419484-5034-4a0f-a2ac-d3f2e3acbaa2",
    "CZ071": "abd57e13-cae2-4514-bc41-bce92351d2c0",
    "CZ072": "40864350-f34a-4f52-a16c-613cb15b73b9",
    "CZ080": "1373e5cf-692d-4c54-901a-67e26ad265b5",
}
_ISPV_REGIONAL_URL_TMPL = (
    "https://www.ispv.cz/getattachment/{guid}"
    "/Vysledky-ve-formatu-XLS.aspx?disposition=attachment"
)
_ISPV_NAT_2025_URL = (
    "https://www.ispv.cz/getattachment/b568f503-6978-4af7-9f8a-d5aef8e46619"
    "/CR_254_MZS-xlsx.aspx?disposition=attachment"
)
_NUTS3_TO_REGION: dict[str, str] = {
    "CZ010": "Hlavní město Praha",
    "CZ020": "Středočeský kraj",
    "CZ031": "Jihočeský kraj",
    "CZ032": "Plzeňský kraj",
    "CZ041": "Karlovarský kraj",
    "CZ042": "Ústecký kraj",
    "CZ051": "Liberecký kraj",
    "CZ052": "Královéhradecký kraj",
    "CZ053": "Pardubický kraj",
    "CZ063": "Kraj Vysočina",
    "CZ064": "Jihomoravský kraj",
    "CZ071": "Olomoucký kraj",
    "CZ072": "Zlínský kraj",
    "CZ080": "Moravskoslezský kraj",
}


# ════════════════════════════════════════════════════════════════════════════
# ISPV Excel parsers
# ════════════════════════════════════════════════════════════════════════════

def _find_sheet(path: Path, keywords: list[str]) -> int | str | None:
    """Return the first sheet name/index that contains one of the keywords."""
    try:
        import openpyxl
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        for i, name in enumerate(wb.sheetnames):
            if any(kw.lower() in name.lower() for kw in keywords):
                wb.close()
                return name
        wb.close()
        # Fallback: return second sheet if available (often regional data)
        return None
    except Exception as exc:
        log.debug("_find_sheet failed: %s", exc)
        return None


def _read_ispv_table(
    path: Path,
    sheet: int | str = 0,
    stat_col_keywords: list[str] | None = None,
) -> pd.DataFrame | None:
    """Read an ISPV Excel sheet; return a DataFrame with sector rows.

    Tries up to 8 skiprows values until a table with >= 3 numeric rows and
    >= 2 columns is found.  Returns None when all attempts fail.

    Parameters
    ----------
    stat_col_keywords:
        Substrings used to locate value columns.  When None, all numeric
        columns are kept.
    """
    for skiprows in range(0, 8):
        try:
            df = pd.read_excel(path, sheet_name=sheet, skiprows=skiprows, header=0)
            df = df.dropna(how="all").reset_index(drop=True)
            if df.shape[1] < 2 or df.shape[0] < 3:
                continue
            first_col = df.columns[0]
            mask = df[first_col].notna() & (df[first_col].astype(str).str.strip() != "")
            sub = df.loc[mask].copy()
            if sub.shape[0] < 3:
                continue
            # Confirm there are at least some numeric values
            numeric_cols = [
                c for c in sub.columns[1:]
                if pd.to_numeric(sub[c], errors="coerce").notna().sum() >= 2
            ]
            if len(numeric_cols) < 1:
                continue
            if stat_col_keywords:
                matched = [
                    c for c in numeric_cols
                    if any(kw.lower() in str(c).lower() for kw in stat_col_keywords)
                ]
                if matched:
                    return sub[[first_col] + matched]
            return sub[[first_col] + numeric_cols]
        except Exception as exc:
            log.debug("_read_ispv_table sheet=%s skiprows=%d: %s", sheet, skiprows, exc)
    return None


def _parse_regional(path: Path) -> pd.Series | None:
    """Return median wage indexed by region label from the ISPV regional sheet."""
    sheet = _find_sheet(path, ["kraj", "region", "územ"])
    if sheet is None:
        sheet = 1  # 2nd sheet is often regional in ISPV workbooks
    df = _read_ispv_table(path, sheet=sheet,
                          stat_col_keywords=["medián", "median", "střední"])
    if df is None:
        return None
    first_col = df.columns[0]
    val_col = df.columns[1]
    s = pd.to_numeric(df[val_col], errors="coerce")
    df = df.copy()
    df[val_col] = s
    df = df.dropna(subset=[val_col])
    df = df[(df[val_col] > 1_000) & (df[val_col] < 500_000)]
    if df.empty:
        return None
    result = df.set_index(first_col)[val_col]
    return result


def _parse_gender_sector(path: Path) -> pd.DataFrame | None:
    """Return a DataFrame with columns [sector, male_median, female_median].

    Looks for a sheet with sex-disaggregated data (columns containing 'muž'
    or 'žen' / 'M' / 'Ž') alongside sector rows.
    """
    male_kws   = ["muž", "muže", "muži", "men", " m "]
    female_kws = ["žen", "ženy", "women", " ž "]
    # Try each sheet
    try:
        import openpyxl
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        sheets = wb.sheetnames
        wb.close()
    except Exception:
        sheets = [0]

    for sheet in sheets:
        df = _read_ispv_table(path, sheet=sheet)
        if df is None:
            continue
        first_col = df.columns[0]
        col_strs = [str(c).lower() for c in df.columns[1:]]
        male_cols   = [df.columns[1 + i] for i, cs in enumerate(col_strs)
                       if any(kw in cs for kw in male_kws)]
        female_cols = [df.columns[1 + i] for i, cs in enumerate(col_strs)
                       if any(kw in cs for kw in female_kws)]
        if not male_cols or not female_cols:
            continue
        m_col = male_cols[0]
        f_col = female_cols[0]
        out = pd.DataFrame({
            "sector":        df[first_col].values,
            "male_median":   pd.to_numeric(df[m_col], errors="coerce").values,
            "female_median": pd.to_numeric(df[f_col], errors="coerce").values,
        }).dropna()
        out = out[(out["male_median"] > 1_000) & (out["female_median"] > 1_000)]
        if len(out) >= 4:
            out["gpg_pct"] = ((out["male_median"] - out["female_median"])
                              / out["male_median"] * 100)
            return out
    return None


def _parse_mzs_percentile_sheet(path: Path, sheet: str) -> pd.DataFrame | None:
    """Parse a structured MZS-style sheet (multi-row header, fixed column order).

    Column layout (after skiprows=7):
      col0=code, col1=label, col2=count, col3=median, col4=YoY%,
      col5=P10, col6=P25, col7=P75, col8=P90, col9=avg, ...
    """
    try:
        df = pd.read_excel(path, sheet_name=sheet, skiprows=7, header=None)
        df = df.dropna(how="all").reset_index(drop=True)
        if df.shape[1] < 8:
            return None
        # Col 1 = label (string), col 3 = median, col 5=P10, col 6=P25, col 7=P75, col 8=P90
        label_col = 1
        col_map = {"p50": 3, "p10": 5, "p25": 6, "p75": 7, "p90": 8}
        # Filter rows where label is a non-null string and median is a plausible wage
        df[label_col] = df[label_col].astype(str).str.strip()
        df = df[df[label_col].str.len() > 2]
        for c in col_map.values():
            df[c] = pd.to_numeric(df[c], errors="coerce")
        df = df.dropna(subset=[col_map["p50"]])
        df = df[df[col_map["p50"]] > 1_000]
        if len(df) < 3:
            return None
        out = pd.DataFrame({"sector": df[label_col].values})
        for pname, col in col_map.items():
            out[pname] = df[col].values
        return out
    except Exception as exc:
        log.debug("_parse_mzs_percentile_sheet %s: %s", sheet, exc)
        return None


def _parse_percentiles(path: Path) -> pd.DataFrame | None:
    """Return sector-level P25/P50/P75 (and optionally P10/P90) from ISPV."""
    # First try structured MZS sheets (MZS-M7 = ISCO occupation breakdown)
    for sheet_try in ["MZS-M7", "MZS-M4", "MZS-M3"]:
        try:
            import openpyxl
            wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
            has_sheet = sheet_try in wb.sheetnames
            wb.close()
        except Exception:
            has_sheet = False
        if has_sheet:
            result = _parse_mzs_percentile_sheet(path, sheet_try)
            if result is not None and len(result) >= 3:
                return result

    # Fallback: heuristic column-matching (old format)
    p_kws = ["p10", "p25", "p50", "p75", "p90", "medián", "median",
             "1. decil", "1. kvartil", "medián", "3. kvartil", "9. decil"]
    df = _read_ispv_table(path, sheet=0, stat_col_keywords=p_kws)
    if df is None:
        return None
    first_col = df.columns[0]
    # Identify P25, P50, P75 columns by heuristic name matching
    percentile_map: dict[str, str] = {}
    for col in df.columns[1:]:
        cs = str(col).lower()
        if "p25" in cs or "1. kvartil" in cs or "25" in cs:
            percentile_map.setdefault("p25", col)
        elif "p75" in cs or "3. kvartil" in cs or "75" in cs:
            percentile_map.setdefault("p75", col)
        elif "p50" in cs or "medián" in cs or "median" in cs or "50" in cs:
            percentile_map.setdefault("p50", col)
        elif "p10" in cs or "1. decil" in cs or "10" in cs:
            percentile_map.setdefault("p10", col)
        elif "p90" in cs or "9. decil" in cs or "90" in cs:
            percentile_map.setdefault("p90", col)
    if len(percentile_map) < 2:
        return None
    keep = [first_col] + list(percentile_map.values())
    sub = df[keep].copy()
    for c in keep[1:]:
        sub[c] = pd.to_numeric(sub[c], errors="coerce")
    sub = sub.dropna()
    sub = sub[(sub.iloc[:, 1] > 1_000)]
    if sub.empty:
        return None
    sub = sub.rename(columns={v: k for k, v in percentile_map.items()})
    sub = sub.rename(columns={first_col: "sector"})
    return sub


def _extract_mzs_m0_median(path: Path) -> float | None:
    """Extract aggregate median wage from the MZS-M0 summary sheet layout.

    The sheet has a sparse layout: each metric is on a separate row with text
    in one cell and the numeric value in the next non-None cell in the same row.
    """
    try:
        import openpyxl
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        sheet = "MZS-M0" if "MZS-M0" in wb.sheetnames else wb.sheetnames[0]
        ws = wb[sheet]
        for row in ws.iter_rows(values_only=True):
            nonnull = [(i, v) for i, v in enumerate(row) if v is not None]
            for idx, (ci, v) in enumerate(nonnull):
                if isinstance(v, str) and "medián" in v.lower() and idx + 1 < len(nonnull):
                    nxt = nonnull[idx + 1][1]
                    if isinstance(nxt, (int, float)) and 10_000 < nxt < 200_000:
                        wb.close()
                        return float(nxt)
        wb.close()
    except Exception as exc:
        log.debug("_extract_mzs_m0_median failed: %s", exc)
    return None


def _fetch_ispv_regional_medians() -> dict[str, float]:
    """Download 14 regional ISPV workbooks; return {nuts3: median_wage_kc}."""
    result: dict[str, float] = {}
    for nuts3, guid in _ISPV_REGIONAL_GUIDS.items():
        url = _ISPV_REGIONAL_URL_TMPL.format(guid=guid)
        try:
            p = fetch(url, suffix=".xlsx")
            with open(p, "rb") as _fh:
                if _fh.read(2) != b"PK":
                    log.debug("%s: not a valid XLSX", nuts3)
                    continue
            med = _extract_mzs_m0_median(p)
            if med is not None:
                result[nuts3] = med
                print(f"    {nuts3}: {med:,.0f} Kč")
        except Exception as exc:
            log.debug("ISPV regional %s failed: %s", nuts3, exc)
    return result


# ════════════════════════════════════════════════════════════════════════════
# Fetch ISPV for latest available year
# ════════════════════════════════════════════════════════════════════════════
print("Fetching ISPV data …")
ispv_path: Path | None = None
ispv_year: int | None = None

for yr in range(END_YEAR, 2014, -1):
    try:
        path_try = fetch_ispv(yr, half=2, sphere="podnikatelska")
        # Quick sanity: at least parseable as Excel
        pd.read_excel(path_try, sheet_name=0, nrows=5)
        ispv_path = path_try
        ispv_year = yr
        print(f"  ISPV {yr}/H2 fetched: {path_try.name}")
        break
    except Exception as exc:
        print(f"  ISPV {yr}/H2: skipped ({type(exc).__name__})")

# Fallback: 2025 national MZS workbook via GUID
if ispv_path is None:
    try:
        p = fetch(_ISPV_NAT_2025_URL, suffix=".xlsx")
        with open(p, "rb") as _fh:
            if _fh.read(2) == b"PK":
                ispv_path, ispv_year = p, 2025
                print(f"  ISPV 2025 GUID fallback: {p.name}")
            else:
                print("  ISPV 2025 GUID fallback: file is not a valid XLSX")
    except Exception as exc:
        print(f"  ISPV 2025 GUID fallback failed: {exc}")

# ════════════════════════════════════════════════════════════════════════════
# Figure A – Regional wage levels
# ════════════════════════════════════════════════════════════════════════════
regional_done = False

if ispv_path is not None:
    reg_wages = _parse_regional(ispv_path)
    if reg_wages is not None and len(reg_wages) >= 4:
        # Identify national average row (often "ČR celkem", "Česká republika")
        avg_kws = ["česká republika", "čr celkem", "celkem", "průměr", "total"]
        nat_row = None
        for idx in reg_wages.index:
            if any(kw in str(idx).lower() for kw in avg_kws):
                nat_row = idx
                break
        if nat_row is not None and reg_wages[nat_row] > 0:
            nat_med = reg_wages[nat_row]
            reg_idx = (reg_wages.drop(index=nat_row) / nat_med * 100).sort_values()
        else:
            nat_med = float(reg_wages.median())
            reg_idx = (reg_wages / nat_med * 100).sort_values()

        # Shorten labels where possible
        reg_idx.index = [_REGION_SHORT.get(str(k), str(k)) for k in reg_idx.index]

        fig_a, ax_a = plt.subplots(figsize=cm2in(16, max(9, len(reg_idx) * 0.65)))
        bar_colors = [CZ_COLOR if v >= 100 else "#4393C3" for v in reg_idx]
        ax_a.barh(reg_idx.index, reg_idx.values, color=bar_colors, alpha=0.82, height=0.7)
        ax_a.axvline(100, color="gray", linewidth=1.2, linestyle="--", zorder=5)
        ax_a.set_xlabel("Index (národní medián = 100)", fontsize=FONT_SIZE)
        ax_a.set_title(
            f"ČR: mediánová mzda podle kraje (ISPV {ispv_year}/H2)",
            fontsize=FONT_SIZE,
        )
        ax_a.xaxis.set_major_formatter(
            ticker.FuncFormatter(lambda x, _: f"{x:.0f}")
        )
        above = mpatches.Patch(color=CZ_COLOR, alpha=0.82, label="Nadnárodní medián")
        below = mpatches.Patch(color="#4393C3", alpha=0.82, label="Podnárodní medián")
        ax_a.legend(handles=[above, below], frameon=False, fontsize=FONT_SIZE - 1,
                    loc="lower right")
        savefig(fig_a, "rscp_regional_wages", out_dir=LATEX_PICS_DIR)
        save_figure_tex(
            "rscp_regional_wages",
            caption=(
                f"ČR: mediánová hrubá mzda podle kraje (ISPV {ispv_year}/H2, "
                "MPSV/TREXIMA), normovaná na národní medián\u00a0=\u00a0100. "
                "Červené sloupce = regiony s~nadprůměrnou mzdou; modré = "
                "podprůměrné kraje. Přetrvávající regionální mzdové nerovnosti "
                "dokládají, proč celostátní minimální mzda a KS různě ovlivňují "
                "reálnou kupní sílu zaměstnanců v~různých částech republiky."
            ),
            cite_keys="mpsv_ispv",
            label="fig:rscp_regional_wages",
            width=r"0.95\linewidth",
            cite_key="mpsv_ispv",
        )
        regional_done = True

if not regional_done:
    # Fallback: download 14 regional ISPV workbooks (NUTS3)
    print("  ISPV regional data unavailable; trying 14 regional ISPV workbooks …")
    regional_medians = _fetch_ispv_regional_medians()
    if len(regional_medians) >= 4:
        nat_med = float(pd.Series(list(regional_medians.values())).median())
        reg_idx_data = {
            _REGION_SHORT.get(_NUTS3_TO_REGION.get(k, k), _NUTS3_TO_REGION.get(k, k)):
            v / nat_med * 100
            for k, v in regional_medians.items()
        }
        s_idx = pd.Series(reg_idx_data).sort_values()
        fig_a2, ax_a2 = plt.subplots(figsize=cm2in(16, max(9, len(s_idx) * 0.65)))
        bar_colors2 = [CZ_COLOR if v >= 100 else "#4393C3" for v in s_idx.values]
        ax_a2.barh(s_idx.index, s_idx.values, color=bar_colors2, alpha=0.82, height=0.7)
        ax_a2.axvline(100, color="gray", linewidth=1.2, linestyle="--", zorder=5)
        ax_a2.set_xlabel("Index (národní medián = 100)", fontsize=FONT_SIZE)
        ax_a2.set_title(
            "ČR: mediánová mzda podle kraje (ISPV 2025/H1, MPSV/TREXIMA)",
            fontsize=FONT_SIZE,
        )
        ax_a2.xaxis.set_major_formatter(
            ticker.FuncFormatter(lambda x, _: f"{x:.0f}")
        )
        above = mpatches.Patch(color=CZ_COLOR, alpha=0.82, label="Nadnárodní medián")
        below = mpatches.Patch(color="#4393C3", alpha=0.82, label="Podnárodní medián")
        ax_a2.legend(handles=[above, below], frameon=False, fontsize=FONT_SIZE - 1,
                     loc="lower right")
        savefig(fig_a2, "rscp_regional_wages", out_dir=LATEX_PICS_DIR)
        save_figure_tex(
            "rscp_regional_wages",
            caption=(
                "ČR: mediánová hrubá mzda podle kraje (ISPV 2025/H1, MPSV/TREXIMA), "
                "normovaná na národní medián\u00a0=\u00a0100. "
                "Červené sloupce\u00a0= regiony s~nadprůměrnou mzdou; "
                "modré\u00a0= podprůměrné kraje. "
                "Přetrvávající regionální mzdové nerovnosti dokládají, "
                "proč celostátní minimální mzda a KS různě ovlivňují "
                "reálnou kupní sílu zaměstnanců v~různých částech republiky."
            ),
            label="fig:rscp_regional_wages",
            width=r"0.95\linewidth",
            cite_key="mpsv_ispv",
        )
        regional_done = True
    else:
        print(f"  Only {len(regional_medians)} ISPV regional files parsed; skipping.")

if not regional_done:
    print("Figure A (regional wages) skipped – no data available.")


# ════════════════════════════════════════════════════════════════════════════
# Figure B – Gender pay gap (sector CZ + cross-country trend)
# ════════════════════════════════════════════════════════════════════════════
print("\nBuilding gender pay gap figures …")

# ── B1: CZ GPG by sector from ISPV ───────────────────────────────────────
ispv_gpg_sector: pd.DataFrame | None = None
if ispv_path is not None:
    ispv_gpg_sector = _parse_gender_sector(ispv_path)
    if ispv_gpg_sector is not None:
        print(f"  ISPV gender breakdown: {len(ispv_gpg_sector)} sectors")
    else:
        print("  ISPV gender columns not found in workbook")

# ── B2: Cross-country GPG trend (Eurostat earn_gr_gpgr2) ─────────────────
geo_6 = "+".join(COUNTRIES)
gpg_trend: dict[str, pd.Series] = {}
try:
    path_gpg = fetch_eurostat(
        "earn_gr_gpgr2",
        f"A.PC.B-S_X_O.{geo_6}",
        start_period=START_YEAR,
    )
    from stattool.dataset import Dataset
    ds_gpg = Dataset.from_sdmx_csv(path_gpg, name="GPG", unit="%",
                                   source_url="Eurostat/earn_gr_gpgr2")
    if not ds_gpg.df.empty:
        for country in COUNTRIES:
            sub = ds_gpg.df[ds_gpg.df["geo"] == country].copy()
            sub["time"] = sub[ds_gpg.time_col].astype(int)
            sub = sub.sort_values("time")
            s = sub.set_index("time")[ds_gpg.value_col].dropna()
            if len(s) >= 3:
                gpg_trend[country] = s
        print(f"  GPG trend countries: {sorted(gpg_trend.keys())}")
    else:
        print("  earn_gr_gpgr2 returned no data")
except Exception as exc:
    print(f"  earn_gr_gpgr2 fetch failed: {exc}")

# ── Build Figure B ────────────────────────────────────────────────────────
has_sector_gpg = ispv_gpg_sector is not None
has_trend_gpg  = bool(gpg_trend)

if has_sector_gpg or has_trend_gpg:
    n_panels = (1 if has_sector_gpg else 0) + (1 if has_trend_gpg else 0)
    fig_b, axes_b = plt.subplots(
        1, n_panels, figsize=cm2in(16 * n_panels / 2 + 4, 10),
        squeeze=False,
    )
    axes_b = axes_b.flatten()
    panel = 0

    if has_sector_gpg:
        ax = axes_b[panel]; panel += 1
        gpg_s = ispv_gpg_sector.sort_values("gpg_pct")
        bar_clrs = [CZ_COLOR if v >= 0 else "#2CA02C" for v in gpg_s["gpg_pct"]]
        ax.barh(gpg_s["sector"].values, gpg_s["gpg_pct"].values,
                color=bar_clrs, alpha=0.82, height=0.7)
        ax.axvline(0, color="gray", linewidth=0.8, linestyle="-")
        ax.set_xlabel("Gender pay gap, muži − ženy (%)", fontsize=FONT_SIZE)
        ax.set_title(
            f"ČR: gender pay gap podle odvětví\n(ISPV {ispv_year}/H2)",
            fontsize=FONT_SIZE,
        )
        ax.xaxis.set_major_formatter(
            ticker.FuncFormatter(lambda x, _: f"{x:.0f}\u00a0pp")
        )

    if has_trend_gpg:
        ax = axes_b[panel]; panel += 1
        for i, country in enumerate(COUNTRIES):
            if country not in gpg_trend:
                continue
            s = gpg_trend[country]
            lw  = 2.5 if country == "CZ" else 1.4
            ls  = "-" if country == "CZ" else "--"
            ax.plot(
                s.index, s.values,
                label=country, color=COUNTRY_COLORS.get(country, PALETTE[i]),
                linewidth=lw, linestyle=ls,
                marker="o" if country == "CZ" else None,
                markersize=3.5,
            )
        ax.yaxis.set_major_formatter(
            ticker.FuncFormatter(lambda y, _: f"{y:.0f}\u00a0%")
        )
        ax.xaxis.set_major_locator(ticker.MaxNLocator(integer=True, nbins=7))
        ax.set_xlabel("rok", fontsize=FONT_SIZE)
        ax.set_ylabel("Neupr. gender pay gap (%)", fontsize=FONT_SIZE)
        ax.set_title(
            "Nekorigovaný gender pay gap: 6 zemí",
            fontsize=FONT_SIZE,
        )
        ax.legend(frameon=False, fontsize=FONT_SIZE - 1.5, ncol=2)

    fig_b.tight_layout(pad=1.5)
    savefig(fig_b, "rscp_gender_gap", out_dir=LATEX_PICS_DIR, tight=False)
    caption_gpg = (
        "Gender pay gap (GPG, rozdíl mediánových hrubých mezd mužů a žen v~\\%) "
    )
    if has_sector_gpg:
        caption_gpg += (
            f"podle odvětví NACE v~ČR (ISPV {ispv_year}/H2, levý panel) a "
        )
    caption_gpg += (
        "vývoj nekorigovaného GPG v~6 srovnávaných zemích "
        f"{START_YEAR}--{END_YEAR} (Eurostat earn_gr_gpgr2, pravý panel). "
        "ČR se dlouhodobě nachází nad průměrem EU a odvětvový profil GPG "
        "ukazuje, že odvětví se silnějším kolektivním vyjednáváním vykazují "
        "nižší mzdové nerovnosti mezi pohlavími."
    )
    save_figure_tex(
        "rscp_gender_gap",
        caption=caption_gpg,
        label="fig:rscp_gender_gap",
        width=r"0.98\linewidth",
        cite_key="eurostat_gpg",
    )
else:
    print("Figure B (gender pay gap) skipped – no data available.")


# ════════════════════════════════════════════════════════════════════════════
# Figure C – Sector wage distribution (P25 / P50 / P75) from ISPV
# ════════════════════════════════════════════════════════════════════════════
print("\nBuilding sector percentile figure …")
pct_df: pd.DataFrame | None = None
if ispv_path is not None:
    pct_df = _parse_percentiles(ispv_path)
    if pct_df is not None:
        print(f"  Percentile table: {len(pct_df)} sectors, cols={list(pct_df.columns)}")
    else:
        print("  No percentile columns found in ISPV workbook")

if pct_df is not None and "p50" in pct_df.columns:
    # Sort by median descending
    pct_df = pct_df.sort_values("p50", ascending=True).reset_index(drop=True)
    n = len(pct_df)

    fig_c, ax_c = plt.subplots(figsize=cm2in(16, max(9, n * 0.65)))

    y_pos = np.arange(n)
    labels = [str(s) for s in pct_df["sector"]]

    # Draw IQR bar (P25–P75) if available
    if "p25" in pct_df.columns and "p75" in pct_df.columns:
        widths = pct_df["p75"].values - pct_df["p25"].values
        ax_c.barh(
            y_pos, widths, left=pct_df["p25"].values,
            color="#4393C3", alpha=0.55, height=0.6, label="P25–P75 (IQR)",
        )

    # P50 tick mark
    ax_c.scatter(
        pct_df["p50"].values, y_pos,
        color=CZ_COLOR, zorder=5, s=40, label="Medián (P50)",
    )

    # P10 / P90 whiskers if available
    if "p10" in pct_df.columns:
        for i, (p10, p25) in enumerate(zip(pct_df["p10"].values,
                                           pct_df.get("p25", pct_df["p50"]).values)):
            ax_c.plot([p10, p25], [i, i], color="gray", linewidth=1.0, alpha=0.7)
    if "p90" in pct_df.columns:
        p75_vals = pct_df.get("p75", pct_df["p50"]).values
        for i, (p75, p90) in enumerate(zip(p75_vals, pct_df["p90"].values)):
            ax_c.plot([p75, p90], [i, i], color="gray", linewidth=1.0, alpha=0.7)

    ax_c.set_yticks(y_pos)
    ax_c.set_yticklabels(labels, fontsize=FONT_SIZE - 1)
    ax_c.xaxis.set_major_formatter(
        ticker.FuncFormatter(lambda x, _: f"{x/1_000:.0f}\u00a0tis. Kč")
    )
    ax_c.set_xlabel("Hrubá měsíční mzda (Kč)", fontsize=FONT_SIZE)
    ax_c.set_title(
        f"ČR: rozložení mezd podle odvětví NACE (ISPV {ispv_year}/H2)\n"
        "mezikvartilový rozsah P25–P75 a medián",
        fontsize=FONT_SIZE,
    )
    ax_c.legend(frameon=False, fontsize=FONT_SIZE - 1, loc="lower right")

    savefig(fig_c, "rscp_sector_percentiles", out_dir=LATEX_PICS_DIR)
    save_figure_tex(
        "rscp_sector_percentiles",
        caption=(
            f"Mzdové rozdělení podle odvětví (CZ-ISCO), ČR, {ispv_year}.okým IQR vyžadují "
            "flexibilnější přístupy nebo individuální sjednávání mzdy."
        ),
        cite_keys="mpsv_ispv",
    label="fig:rscp_sector_percentiles",
        width=r"0.95\linewidth",
        cite_key="mpsv_ispv",
    )
else:
    print("Figure C (sector percentiles) skipped – no percentile data available.")


# ════════════════════════════════════════════════════════════════════════════
# Figure D – CZ NUTS3 regional wage choropleth (ISPV 14 regional workbooks)
# ════════════════════════════════════════════════════════════════════════════
print("\nBuilding CZ NUTS3 regional wage choropleth (ISPV) …")

import geopandas as gpd
from shapely.geometry import box as _shapely_box

_GISCO_NUTS_URL = (
    "https://gisco-services.ec.europa.eu/distribution/v2/nuts/geojson/"
    "NUTS_RG_20M_2021_3035.geojson"
)
# EPSG:3035 bounding box for CZ and immediate neighbours (in metres)
_CZ_XLIM = (4_300_000, 5_100_000)
_CZ_YLIM = (2_700_000, 3_300_000)

try:
    # ── 1. Download NUTS3 GeoJSON (cached) ───────────────────────────────
    nuts_path = fetch(_GISCO_NUTS_URL, suffix=".geojson")
    nuts = gpd.read_file(nuts_path)
    nuts3 = nuts[nuts["LEVL_CODE"] == 3].copy()
    nuts3["CNTR_CODE"] = nuts3["NUTS_ID"].str[:2]

    # ── 2. Download 14 regional ISPV workbooks ────────────────────────────
    print("  Fetching 14 regional ISPV medians …")
    reg_medians = _fetch_ispv_regional_medians()
    if len(reg_medians) < 4:
        raise ValueError(f"Only {len(reg_medians)} ISPV regional files parseable")
    nat_avg = float(pd.Series(list(reg_medians.values())).median())
    wage_df = pd.DataFrame(
        [{"NUTS_ID": k, "wage_idx": v / nat_avg * 100}
         for k, v in reg_medians.items()]
    )

    # ── 3. Join to geometry ───────────────────────────────────────────────
    merged = nuts3.merge(wage_df, on="NUTS_ID", how="left")

    # ── 4. Draw map ───────────────────────────────────────────────────────
    import matplotlib.colors as mcolors
    import matplotlib.cm as mcm

    fig_d, ax_d = plt.subplots(1, 1, figsize=cm2in(15, 11))

    # Grey basemap for neighbours (NUTS3 of AT, DE, PL, SK)
    neighbours = merged[merged["CNTR_CODE"].isin(["AT", "DE", "PL", "SK"])]
    neighbours.plot(ax=ax_d, color="#DDDDDD", edgecolor="white", linewidth=0.4)

    # CZ regions coloured by wage index
    cz_regions = merged[merged["CNTR_CODE"] == "CZ"].copy()
    bbox = _shapely_box(_CZ_XLIM[0], _CZ_YLIM[0], _CZ_XLIM[1], _CZ_YLIM[1])
    cz_regions = cz_regions.copy()
    cz_regions["geometry"] = cz_regions["geometry"].intersection(bbox)

    cmap = plt.colormaps["RdYlGn"]
    vmin, vmax = 70, 150
    norm = mcolors.Normalize(vmin=vmin, vmax=vmax)
    for _, row in cz_regions.iterrows():
        if row.geometry is None or row.geometry.is_empty:
            continue
        val = row["wage_idx"]
        color = cmap(norm(val)) if pd.notna(val) else "#CCCCCC"
        gpd.GeoSeries([row.geometry]).plot(
            ax=ax_d, color=color, edgecolor="white", linewidth=0.5,
        )
        centroid = row.geometry.centroid
        if pd.notna(val):
            ax_d.text(
                centroid.x, centroid.y,
                f"{val:.0f}",
                ha="center", va="center",
                fontsize=FONT_SIZE - 2.5, color="black",
            )

    # Colourbar
    sm = mcm.ScalarMappable(cmap=cmap, norm=norm)
    sm.set_array([])
    cbar = fig_d.colorbar(sm, ax=ax_d, fraction=0.03, pad=0.02)
    cbar.set_label("Index (medián ČR = 100)", fontsize=FONT_SIZE)
    cbar.ax.tick_params(labelsize=FONT_SIZE - 1)

    ax_d.set_xlim(_CZ_XLIM)
    ax_d.set_ylim(_CZ_YLIM)
    ax_d.axis("off")
    ax_d.set_title(
        "ČR: mediánová mzda podle kraje (ISPV 2025/H1)\n"
        "index (medián ČR = 100)",
        fontsize=FONT_SIZE,
    )
    fig_d.tight_layout()

    savefig(fig_d, "rscp_regional_wages_map", out_dir=LATEX_PICS_DIR)
    save_figure_tex(
        "rscp_regional_wages_map",
        caption=(
            "ČR: mediánová hrubá mzda podle kraje (NUTS3) (ISPV 2025/H1, "
            "MPSV/TREXIMA), normovaná na národní medián\u00a0=\u00a0100. "
            "Sousední regiony (AT, DE, PL, SK) zobrazeny šedě. "
            "Praha (CZ010) a přilehlý Středočeský kraj (CZ020) dosahují "
            "výrazně vyšších mezd než periferní regiony, "
            "což zesiluje emigrační tlak a ztěžuje praktický dopad "
            "celostátních kolektivních smluv."
        ),
        cite_keys="mpsv_ispv",
    label="fig:rscp_regional_wages_map",
        width=r"0.85\linewidth",
        cite_key="mpsv_ispv",
    )
    print("  Figure D done.")

except Exception as exc:
    print(f"Figure D (NUTS3 choropleth) skipped: {exc}")


print("\nDone.")
